"""
Generador del archivo "Sistema_Gestion_Financiera.xlsx"
Mini ERP para empresas venezolanas - Fase 1 (Core financiero)
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import date

# ---------------------------------------------------------------
# TEMA CORPORATIVO
# ---------------------------------------------------------------
NAVY = "1F4E78"
BLUE = "2E75B6"
LIGHT_BLUE = "DDEBF7"
GREEN = "548235"
RED = "C00000"
YELLOW = "BF8F00"
GREY = "F2F2F2"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Segoe UI", size=20, bold=True, color=WHITE)
H1_FONT = Font(name="Segoe UI", size=14, bold=True, color=WHITE)
H2_FONT = Font(name="Segoe UI", size=11, bold=True, color=NAVY)
LABEL_FONT = Font(name="Segoe UI", size=10, bold=True)
NORMAL_FONT = Font(name="Segoe UI", size=10)
KPI_FONT = Font(name="Segoe UI", size=18, bold=True, color=NAVY)
KPI_LABEL_FONT = Font(name="Segoe UI", size=9, bold=True, color=WHITE)

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SUBHEADER_FILL = PatternFill("solid", fgColor=BLUE)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
KPI_FILL = PatternFill("solid", fgColor=BLUE)
GREY_FILL = PatternFill("solid", fgColor=GREY)

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

CURRENCY_BS = '#,##0.00 "Bs"'
CURRENCY_USD = '$#,##0.00'
PCT = '0.0%'

wb = Workbook()


def style_header_row(ws, row, max_col, fill=HEADER_FILL, font=H1_FONT):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------
# 01_PORTADA
# ---------------------------------------------------------------
ws = wb.active
ws.title = "01_PORTADA"
ws.sheet_view.showGridLines = False

ws.merge_cells("B2:H6")
c = ws["B2"]
c.value = "SISTEMA INTEGRAL DE GESTION FINANCIERA"
c.font = Font(name="Segoe UI", size=26, bold=True, color=NAVY)
c.alignment = CENTER

ws.merge_cells("B7:H7")
c = ws["B7"]
c.value = "[NOMBRE DE LA EMPRESA]"
c.font = Font(name="Segoe UI", size=16, bold=True, color=BLUE)
c.alignment = CENTER

for r in range(2, 8):
    for col in range(2, 9):
        ws.cell(row=r, column=col).fill = LIGHT_FILL

info = [
    ("RIF:", "J-00000000-0"),
    ("Direccion:", "[Direccion fiscal]"),
    ("Ano fiscal:", date.today().year),
    ("Moneda principal:", "Bs / USD"),
    ("Fecha de actualizacion:", date.today()),
]
r = 9
for label, val in info:
    ws[f"B{r}"] = label
    ws[f"B{r}"].font = LABEL_FONT
    ws[f"D{r}"] = val
    ws[f"D{r}"].font = NORMAL_FONT
    r += 1

# Menu de navegacion
ws[f"B{r+1}"] = "MENU DE NAVEGACION"
ws[f"B{r+1}"].font = H1_FONT
ws[f"B{r+1}"].fill = HEADER_FILL
ws.merge_cells(start_row=r+1, start_column=2, end_row=r+1, end_column=4)

menu = [
    "00_MANUAL", "02_CONFIGURACION", "03_DASHBOARD_EJECUTIVO", "04_INGRESOS", "05_EGRESOS",
    "06_COSTOS", "07_PUNTO_EQUILIBRIO", "08_FLUJO_CAJA", "09_ESTADO_RESULTADOS",
    "10_BALANCE_GENERAL", "11_CUENTAS_POR_COBRAR", "12_CUENTAS_POR_PAGAR", "13_INVENTARIO",
    "14_NOMINA", "15_IMPUESTOS", "16_PRESUPUESTO", "17_KPI_GERENCIALES",
    "18_PRODUCCION_MINERA", "19_COSTO_POR_TONELADA", "20_RESERVAS_MINERAS", "21_INDICADORES_MINEROS"
]
rr = r + 2
for m in menu:
    cell = ws.cell(row=rr, column=2, value=m)
    cell.font = Font(name="Segoe UI", size=11, color=BLUE, underline="single")
    cell.hyperlink = f"#'{m}'!A1"
    rr += 1

autosize(ws, {"A": 3, "B": 26, "C": 4, "D": 26, "E": 4, "F": 4, "G": 4, "H": 4})

# ---------------------------------------------------------------
# 02_CONFIGURACION
# ---------------------------------------------------------------
ws = wb.create_sheet("02_CONFIGURACION")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:D1")
ws["A1"] = "PARAMETROS GENERALES DE LA EMPRESA"
style_header_row(ws, 1, 4)

params = [
    ("Nombre de la empresa", "[NOMBRE DE LA EMPRESA]"),
    ("RIF", "J-00000000-0"),
    ("Direccion", "[Direccion fiscal]"),
    ("Moneda principal", "Bs"),
    ("Tasa BCV (Bs/USD)", 36.50),
    ("Alicuota IVA general", 0.16),
    ("Alicuota IGTF", 0.03),
    ("Ano fiscal", date.today().year),
]
r = 3
for label, val in params:
    ws.cell(row=r, column=1, value=label).font = LABEL_FONT
    cell = ws.cell(row=r, column=2, value=val)
    cell.fill = LIGHT_FILL
    cell.border = BORDER
    if label == "Tasa BCV (Bs/USD)":
        cell.number_format = CURRENCY_BS
        ws.defined_names.add(openpyxl.workbook.defined_name.DefinedName("Tasa_BCV", attr_text=f"'02_CONFIGURACION'!$B${r}"))
    if "Alicuota" in label:
        cell.number_format = PCT
    r += 1

r += 1
ws.cell(row=r, column=1, value="CENTROS DE COSTOS").font = H2_FONT
r += 1
cc_start = r
for cc in ["Administracion", "Ventas", "Produccion", "Logistica", "Mina"]:
    ws.cell(row=r, column=1, value=cc).fill = LIGHT_FILL
    ws.cell(row=r, column=1).border = BORDER
    r += 1
cc_end = r - 1

r += 1
ws.cell(row=r, column=1, value="LINEAS DE NEGOCIO / PROYECTOS").font = H2_FONT
r += 1
for ln in ["Linea A", "Linea B", "Proyecto Minero 1"]:
    ws.cell(row=r, column=1, value=ln).fill = LIGHT_FILL
    ws.cell(row=r, column=1).border = BORDER
    r += 1

r += 1
ws.cell(row=r, column=1, value="CATEGORIAS DE INGRESOS").font = H2_FONT
r += 1
ing_cat_start = r
for cat in ["Ventas", "Servicios", "Exportaciones", "Otros ingresos"]:
    ws.cell(row=r, column=1, value=cat).fill = LIGHT_FILL
    ws.cell(row=r, column=1).border = BORDER
    r += 1
ing_cat_end = r - 1

r += 1
ws.cell(row=r, column=1, value="CATEGORIAS DE EGRESOS").font = H2_FONT
r += 1
egr_cat_start = r
for cat in ["Nomina", "Combustible", "Transporte", "Mantenimiento", "Servicios",
            "Impuestos", "Compras", "Gastos Financieros", "Otros gastos"]:
    ws.cell(row=r, column=1, value=cat).fill = LIGHT_FILL
    ws.cell(row=r, column=1).border = BORDER
    r += 1
egr_cat_end = r - 1

autosize(ws, {"A": 30, "B": 22, "C": 4, "D": 4})

# ---------------------------------------------------------------
# Helper: build a transactional table sheet (Ingresos/Egresos)
# ---------------------------------------------------------------
def build_transactions_sheet(name, third_party_label, sample_data, table_name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws["A1"] = name.replace("_", " ")
    style_header_row(ws, 1, 7)

    headers = ["Fecha", third_party_label, "Concepto", "Categoria", "Centro de Costos", "Bs.", "USD"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.fill = SUBHEADER_FILL
        cell.font = H1_FONT
        cell.alignment = CENTER

    for ridx, row in enumerate(sample_data, start=4):
        for cidx, val in enumerate(row, start=1):
            cell = ws.cell(row=ridx, column=cidx, value=val)
            cell.border = BORDER
            if cidx in (1,):
                cell.number_format = "dd/mm/yyyy"
            if cidx == 6:
                cell.number_format = CURRENCY_BS
            if cidx == 7:
                cell.number_format = CURRENCY_USD
                cell.value = f"=F{ridx}/Tasa_BCV"

    last_row = 3 + len(sample_data)
    tbl = Table(displayName=table_name, ref=f"A3:G{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    # Totales
    tot_row = last_row + 2
    ws.cell(row=tot_row, column=5, value="TOTAL").font = LABEL_FONT
    ws.cell(row=tot_row, column=6, value=f"=SUBTOTAL(109,{table_name}[Bs.])").number_format = CURRENCY_BS
    ws.cell(row=tot_row, column=7, value=f"=SUBTOTAL(109,{table_name}[USD])").number_format = CURRENCY_USD
    for c in (5, 6, 7):
        ws.cell(row=tot_row, column=c).fill = LIGHT_FILL
        ws.cell(row=tot_row, column=c).border = BORDER
        ws.cell(row=tot_row, column=c).font = LABEL_FONT

    autosize(ws, {"A": 12, "B": 22, "C": 28, "D": 18, "E": 18, "F": 16, "G": 14})
    return ws, table_name, last_row


# ---------------------------------------------------------------
# 04_INGRESOS
# ---------------------------------------------------------------
ingresos_data = [
    (date(2026, 1, 5), "Cliente A", "Venta de producto X", "Ventas", "Ventas", 45000, None),
    (date(2026, 1, 12), "Cliente B", "Servicio de mantenimiento", "Servicios", "Produccion", 18000, None),
    (date(2026, 2, 3), "Cliente C", "Exportacion lote 1", "Exportaciones", "Logistica", 120000, None),
    (date(2026, 2, 20), "Cliente A", "Venta de producto Y", "Ventas", "Ventas", 60000, None),
    (date(2026, 3, 8), "Cliente D", "Otros ingresos varios", "Otros ingresos", "Administracion", 5000, None),
    (date(2026, 3, 15), "Cliente B", "Venta de producto X", "Ventas", "Ventas", 75000, None),
    (date(2026, 4, 2), "Cliente E", "Servicio tecnico", "Servicios", "Produccion", 22000, None),
    (date(2026, 4, 18), "Cliente C", "Exportacion lote 2", "Exportaciones", "Logistica", 135000, None),
    (date(2026, 5, 10), "Cliente A", "Venta de producto Z", "Ventas", "Ventas", 80000, None),
    (date(2026, 5, 25), "Cliente F", "Servicio de consultoria", "Servicios", "Administracion", 30000, None),
    (date(2026, 6, 1), "Cliente B", "Venta mensual", "Ventas", "Ventas", 90000, None),
    (date(2026, 6, 9), "Cliente C", "Exportacion lote 3", "Exportaciones", "Logistica", 150000, None),
]
ws_ing, T_ING, last_ing = build_transactions_sheet("04_INGRESOS", "Cliente", ingresos_data, "tbl_Ingresos")

# ---------------------------------------------------------------
# 05_EGRESOS
# ---------------------------------------------------------------
egresos_data = [
    (date(2026, 1, 5), "Nomina", "Pago quincenal nomina", "Nomina", "Administracion", 30000, None),
    (date(2026, 1, 10), "PDVSA", "Combustible flota", "Combustible", "Logistica", 8000, None),
    (date(2026, 1, 20), "Transportes XYZ", "Flete materiales", "Transporte", "Logistica", 6000, None),
    (date(2026, 2, 5), "Nomina", "Pago quincenal nomina", "Nomina", "Administracion", 30000, None),
    (date(2026, 2, 8), "Taller Mecanico", "Mantenimiento equipos", "Mantenimiento", "Produccion", 9000, None),
    (date(2026, 2, 28), "SENIAT", "IVA del periodo", "Impuestos", "Administracion", 12000, None),
    (date(2026, 3, 5), "Nomina", "Pago quincenal nomina", "Nomina", "Administracion", 31000, None),
    (date(2026, 3, 15), "Proveedor Insumos", "Compra materia prima", "Compras", "Produccion", 40000, None),
    (date(2026, 3, 30), "Banco", "Intereses prestamo", "Gastos Financieros", "Administracion", 4000, None),
    (date(2026, 4, 5), "Nomina", "Pago quincenal nomina", "Nomina", "Administracion", 31000, None),
    (date(2026, 4, 12), "Electricidad CORPOELEC", "Servicio electrico", "Servicios", "Produccion", 7000, None),
    (date(2026, 5, 5), "Nomina", "Pago quincenal nomina", "Nomina", "Administracion", 32000, None),
    (date(2026, 5, 18), "Proveedor Insumos", "Compra materia prima", "Compras", "Produccion", 45000, None),
    (date(2026, 6, 1), "Nomina", "Pago quincenal nomina", "Nomina", "Administracion", 32000, None),
    (date(2026, 6, 8), "PDVSA", "Combustible flota", "Combustible", "Logistica", 9000, None),
]
ws_egr, T_EGR, last_egr = build_transactions_sheet("05_EGRESOS", "Proveedor", egresos_data, "tbl_Egresos")

# ---------------------------------------------------------------
# 06_COSTOS
# ---------------------------------------------------------------
ws = wb.create_sheet("06_COSTOS")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:D1")
ws["A1"] = "ESTRUCTURA DE COSTOS"
style_header_row(ws, 1, 4)

ws["A3"] = "COSTOS FIJOS MENSUALES"
ws["A3"].font = H1_FONT
ws["A3"].fill = SUBHEADER_FILL
ws.merge_cells("A3:B3")

cf = [
    ("Nomina administrativa", 32000),
    ("Seguros", 3000),
    ("Arrendamientos", 8000),
    ("Servicios (luz, agua, internet)", 4000),
]
r = 4
for label, val in cf:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=val)
    cell.number_format = CURRENCY_BS
    cell.border = BORDER
    r += 1
cf_end = r - 1
ws.cell(row=r, column=1, value="TOTAL COSTOS FIJOS").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=SUM(B4:B{cf_end})")
cell.font = LABEL_FONT
cell.fill = LIGHT_FILL
cell.number_format = CURRENCY_BS
cf_total_row = r

r += 2
ws.cell(row=r, column=1, value="COSTOS VARIABLES (por unidad)").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
cv = [
    ("Materia prima", 80),
    ("Transporte", 15),
    ("Combustible", 10),
    ("Comercializacion", 5),
]
cv_start = r
for label, val in cv:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=val)
    cell.number_format = CURRENCY_BS
    cell.border = BORDER
    r += 1
cv_end = r - 1
ws.cell(row=r, column=1, value="COSTO VARIABLE UNITARIO TOTAL").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=SUM(B{cv_start}:B{cv_end})")
cell.font = LABEL_FONT
cell.fill = LIGHT_FILL
cell.number_format = CURRENCY_BS
cvu_row = r

r += 2
ws.cell(row=r, column=1, value="PRECIO DE VENTA UNITARIO").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=180)
cell.number_format = CURRENCY_BS
cell.fill = LIGHT_FILL
pvu_row = r

r += 1
ws.cell(row=r, column=1, value="UNIDADES VENDIDAS (mes)").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=1200)
cell.fill = LIGHT_FILL
unid_row = r

# Named ranges para uso en otras hojas
def add_named(name, ref):
    wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(name, attr_text=ref))

add_named("Costos_Fijos", f"'06_COSTOS'!$B${cf_total_row}")
add_named("Costo_Var_Unitario", f"'06_COSTOS'!$B${cvu_row}")
add_named("Precio_Venta_Unitario", f"'06_COSTOS'!$B${pvu_row}")
add_named("Unidades_Vendidas", f"'06_COSTOS'!$B${unid_row}")

autosize(ws, {"A": 32, "B": 16, "C": 4, "D": 4})

# ---------------------------------------------------------------
# 07_PUNTO_EQUILIBRIO
# ---------------------------------------------------------------
ws = wb.create_sheet("07_PUNTO_EQUILIBRIO")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:D1")
ws["A1"] = "PUNTO DE EQUILIBRIO"
style_header_row(ws, 1, 4)

rows = [
    ("Costos Fijos (Bs.)", "=Costos_Fijos", CURRENCY_BS),
    ("Precio de Venta Unitario (Bs.)", "=Precio_Venta_Unitario", CURRENCY_BS),
    ("Costo Variable Unitario (Bs.)", "=Costo_Var_Unitario", CURRENCY_BS),
    ("Margen de Contribucion Unitario (Bs.)", "=B5-B6", CURRENCY_BS),
    ("Punto de Equilibrio (unidades)", "=B4/B7", '#,##0'),
    ("Punto de Equilibrio (Bs.)", "=B8*B5", CURRENCY_BS),
    ("Punto de Equilibrio (USD)", "=B9/Tasa_BCV", CURRENCY_USD),
    ("Unidades Vendidas Actuales", "=Unidades_Vendidas", '#,##0'),
    ("Margen de Seguridad (unidades)", "=B11-B8", '#,##0'),
    ("Margen de Seguridad (%)", "=B12/B11", PCT),
]
r = 4
for label, formula, fmt in rows:
    ws.cell(row=r, column=1, value=label).font = LABEL_FONT
    cell = ws.cell(row=r, column=2, value=formula)
    cell.fill = LIGHT_FILL
    cell.border = BORDER
    cell.number_format = fmt
    r += 1

# Simulacion de escenarios
r += 1
ws.cell(row=r, column=1, value="SIMULACION DE ESCENARIOS").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
sim_header_row = r
for i, h in enumerate(["Escenario", "Var. Precio (%)", "Var. Costo Var. (%)", "Nuevo PE (unid.)"], start=1):
    cell = ws.cell(row=r, column=i, value=h)
    cell.font = H1_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
r += 1
escenarios = [("Optimista", 0.10, -0.05), ("Base", 0.0, 0.0), ("Pesimista", -0.10, 0.10)]
for nombre, dp, dc in escenarios:
    ws.cell(row=r, column=1, value=nombre).border = BORDER
    cell = ws.cell(row=r, column=2, value=dp); cell.number_format = PCT; cell.border = BORDER
    cell = ws.cell(row=r, column=3, value=dc); cell.number_format = PCT; cell.border = BORDER
    formula = f"=$B$4/(($B$5*(1+B{r}))-($B$6*(1+C{r})))"
    cell = ws.cell(row=r, column=4, value=formula); cell.number_format = '#,##0'; cell.border = BORDER
    r += 1

autosize(ws, {"A": 36, "B": 18, "C": 18, "D": 18})

# ---------------------------------------------------------------
# 08_FLUJO_DE_CAJA
# ---------------------------------------------------------------
ws = wb.create_sheet("08_FLUJO_CAJA")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:E1")
ws["A1"] = "FLUJO DE CAJA MENSUAL"
style_header_row(ws, 1, 5)

headers = ["Mes", "Entradas (Bs.)", "Salidas (Bs.)", "Flujo Neto (Bs.)", "Saldo Acumulado (Bs.)"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = H1_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER

months = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto",
          "Septiembre","Octubre","Noviembre","Diciembre"]
start_row = 4
for i, m in enumerate(months):
    r = start_row + i
    ws.cell(row=r, column=1, value=m).border = BORDER
    f_in = (f"=SUMPRODUCT((MONTH(tbl_Ingresos[Fecha])={i+1})*"
            f"(YEAR(tbl_Ingresos[Fecha])='02_CONFIGURACION'!$B$8)*tbl_Ingresos[Bs.])")
    f_out = (f"=SUMPRODUCT((MONTH(tbl_Egresos[Fecha])={i+1})*"
             f"(YEAR(tbl_Egresos[Fecha])='02_CONFIGURACION'!$B$8)*tbl_Egresos[Bs.])")
    c_in = ws.cell(row=r, column=2, value=f_in); c_in.number_format = CURRENCY_BS; c_in.border = BORDER
    c_out = ws.cell(row=r, column=3, value=f_out); c_out.number_format = CURRENCY_BS; c_out.border = BORDER
    c_net = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); c_net.number_format = CURRENCY_BS; c_net.border = BORDER
    if i == 0:
        c_acc = ws.cell(row=r, column=5, value=f"=D{r}")
    else:
        c_acc = ws.cell(row=r, column=5, value=f"=E{r-1}+D{r}")
    c_acc.number_format = CURRENCY_BS; c_acc.border = BORDER

end_row = start_row + len(months) - 1

# Proyecciones simples (promedio diario * dias)
proj_row = end_row + 2
ws.cell(row=proj_row, column=1, value="PROYECCIONES (a partir del flujo neto promedio mensual)").font = H1_FONT
ws.cell(row=proj_row, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=proj_row, start_column=1, end_row=proj_row, end_column=5)
proj_row += 1
for i, h in enumerate(["Horizonte", "Dias", "Flujo Proyectado (Bs.)", "", ""], start=1):
    if h:
        cell = ws.cell(row=proj_row, column=i, value=h)
        cell.font = H1_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
proj_row += 1
for horizonte, dias in [("30 dias", 30), ("90 dias", 90), ("180 dias", 180), ("365 dias", 365)]:
    ws.cell(row=proj_row, column=1, value=horizonte).border = BORDER
    ws.cell(row=proj_row, column=2, value=dias).border = BORDER
    formula = f"=(AVERAGE(D{start_row}:D{end_row})/30)*B{proj_row}"
    cell = ws.cell(row=proj_row, column=3, value=formula)
    cell.number_format = CURRENCY_BS
    cell.border = BORDER
    proj_row += 1

# Grafico
chart = LineChart()
chart.title = "Flujo de Caja Mensual"
chart.style = 10
data = Reference(ws, min_col=4, min_row=3, max_row=end_row)
cats = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 8
chart.width = 16
ws.add_chart(chart, "G3")

autosize(ws, {"A": 14, "B": 18, "C": 18, "D": 18, "E": 20})

# ---------------------------------------------------------------
# 09_ESTADO_RESULTADOS
# ---------------------------------------------------------------
ws = wb.create_sheet("09_ESTADO_RESULTADOS")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:C1")
ws["A1"] = "ESTADO DE RESULTADOS"
style_header_row(ws, 1, 3)

ing_cats = "Ventas\",\"Servicios\",\"Exportaciones\",\"Otros ingresos"
def sumif_egresos(cat):
    return f'=SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"{cat}")'
def sumif_ingresos_total():
    return ('=SUMIFS(tbl_Ingresos[Bs.],tbl_Ingresos[Categoria],"Ventas")'
            '+SUMIFS(tbl_Ingresos[Bs.],tbl_Ingresos[Categoria],"Servicios")'
            '+SUMIFS(tbl_Ingresos[Bs.],tbl_Ingresos[Categoria],"Exportaciones")'
            '+SUMIFS(tbl_Ingresos[Bs.],tbl_Ingresos[Categoria],"Otros ingresos")')

er_rows = []
er_rows.append(("Ingresos", sumif_ingresos_total(), False))
er_rows.append(("(-) Costo de Ventas (Compras)", sumif_egresos("Compras"), False))
er_rows.append(("= UTILIDAD BRUTA", "=B4-B5", True))
gop = ('=' + '+'.join([
    f'SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"{c}")'
    for c in ["Nomina","Combustible","Transporte","Mantenimiento","Servicios","Otros gastos"]
]))
er_rows.append(("(-) Gastos Operativos", gop, False))
er_rows.append(("= UTILIDAD OPERATIVA (EBIT)", "=B6-B7", True))
er_rows.append(("(-) Gastos Financieros", sumif_egresos("Gastos Financieros"), False))
er_rows.append(("= UTILIDAD ANTES DE IMPUESTOS", "=B8-B9", True))
er_rows.append(("(-) Impuestos", sumif_egresos("Impuestos"), False))
er_rows.append(("= UTILIDAD NETA", "=B10-B11", True))

r = 3
for label, formula, is_total in er_rows:
    cell_a = ws.cell(row=r, column=1, value=label)
    cell_b = ws.cell(row=r, column=2, value=formula)
    cell_b.number_format = CURRENCY_BS
    if is_total:
        cell_a.font = LABEL_FONT
        cell_b.font = LABEL_FONT
        cell_a.fill = LIGHT_FILL
        cell_b.fill = LIGHT_FILL
    else:
        cell_a.font = NORMAL_FONT
        cell_b.font = NORMAL_FONT
    cell_a.border = BORDER
    cell_b.border = BORDER
    r += 1

# Named ranges para Dashboard
add_named("Ingresos_Totales", "='09_ESTADO_RESULTADOS'!$B$3")
add_named("Utilidad_Bruta", "='09_ESTADO_RESULTADOS'!$B$5")
add_named("Utilidad_Operativa", "='09_ESTADO_RESULTADOS'!$B$7")
add_named("Utilidad_Neta", "='09_ESTADO_RESULTADOS'!$B$11")
add_named("Gastos_Financieros", "='09_ESTADO_RESULTADOS'!$B$8")
add_named("Egresos_Totales", f"='05_EGRESOS'!$F${last_egr+2}")

autosize(ws, {"A": 32, "B": 18, "C": 4})

# ---------------------------------------------------------------
# 03_DASHBOARD_EJECUTIVO
# ---------------------------------------------------------------
ws = wb.create_sheet("03_DASHBOARD_EJECUTIVO", 2)  # insertar despues de Configuracion
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:L1")
ws["A1"] = "DASHBOARD EJECUTIVO"
style_header_row(ws, 1, 12)

today_year = date.today().year

# KPI cards: (titulo, formula, formato, columna inicio)
kpis = [
    ("Ingresos del Mes (Bs.)",
     f"=SUMPRODUCT((MONTH(tbl_Ingresos[Fecha])=MONTH(TODAY()))*(YEAR(tbl_Ingresos[Fecha])=YEAR(TODAY()))*tbl_Ingresos[Bs.])",
     CURRENCY_BS, 1),
    ("Ingresos del Ano (Bs.)",
     f"=SUMPRODUCT((YEAR(tbl_Ingresos[Fecha])=YEAR(TODAY()))*tbl_Ingresos[Bs.])",
     CURRENCY_BS, 4),
    ("Egresos del Mes (Bs.)",
     f"=SUMPRODUCT((MONTH(tbl_Egresos[Fecha])=MONTH(TODAY()))*(YEAR(tbl_Egresos[Fecha])=YEAR(TODAY()))*tbl_Egresos[Bs.])",
     CURRENCY_BS, 7),
    ("Egresos del Ano (Bs.)",
     f"=SUMPRODUCT((YEAR(tbl_Egresos[Fecha])=YEAR(TODAY()))*tbl_Egresos[Bs.])",
     CURRENCY_BS, 10),
    ("Utilidad Bruta (Bs.)", "=Utilidad_Bruta", CURRENCY_BS, 1),
    ("Utilidad Operativa (Bs.)", "=Utilidad_Operativa", CURRENCY_BS, 4),
    ("Utilidad Neta (Bs.)", "=Utilidad_Neta", CURRENCY_BS, 7),
    ("EBITDA (Bs.)", "=Utilidad_Operativa", CURRENCY_BS, 10),
    ("Liquidez (veces)", 1.5, '0.00', 1),
    ("Capital de Trabajo (Bs.)", 250000, CURRENCY_BS, 4),
    ("Endeudamiento (%)", 0.45, PCT, 7),
    ("Margen Neto (%)", "=Utilidad_Neta/Ingresos_Totales", PCT, 10),
]

start_row = 3
card_w = 3
for idx, (title, formula, fmt, col) in enumerate(kpis):
    row = start_row + (idx // 4) * 4
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+card_w-1)
    tcell = ws.cell(row=row, column=col, value=title)
    tcell.font = KPI_LABEL_FONT
    tcell.fill = KPI_FILL
    tcell.alignment = CENTER
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+card_w-1)
    vcell = ws.cell(row=row+1, column=col, value=formula)
    vcell.font = KPI_FONT
    vcell.fill = LIGHT_FILL
    vcell.number_format = fmt
    vcell.alignment = CENTER

# fila libre
chart_row = start_row + 3 * 4 + 1

# Grafico de ventas por categoria (Ingresos)
ws.cell(row=chart_row, column=1, value="VENTAS / INGRESOS POR CATEGORIA").font = H2_FONT
cat_table_row = chart_row + 1
cats_list = ["Ventas", "Servicios", "Exportaciones", "Otros ingresos"]
for i, cat in enumerate(cats_list):
    ws.cell(row=cat_table_row + i, column=1, value=cat).border = BORDER
    f = f'=SUMIFS(tbl_Ingresos[Bs.],tbl_Ingresos[Categoria],A{cat_table_row+i})'
    cell = ws.cell(row=cat_table_row + i, column=2, value=f)
    cell.number_format = CURRENCY_BS
    cell.border = BORDER

pie = PieChart()
pie.title = "Ingresos por Categoria"
data = Reference(ws, min_col=2, min_row=cat_table_row, max_row=cat_table_row+len(cats_list)-1)
cats_ref = Reference(ws, min_col=1, min_row=cat_table_row, max_row=cat_table_row+len(cats_list)-1)
pie.add_data(data)
pie.set_categories(cats_ref)
pie.height = 8
pie.width = 10
ws.add_chart(pie, f"D{chart_row}")

# Grafico egresos por categoria
egr_cats_list = ["Nomina","Combustible","Transporte","Mantenimiento","Servicios",
                  "Impuestos","Compras","Gastos Financieros","Otros gastos"]
egr_table_row = cat_table_row
egr_table_col = 9
ws.cell(row=chart_row, column=egr_table_col, value="EGRESOS POR CATEGORIA").font = H2_FONT
for i, cat in enumerate(egr_cats_list):
    ws.cell(row=egr_table_row + i, column=egr_table_col, value=cat).border = BORDER
    f = f'=SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],{get_column_letter(egr_table_col)}{egr_table_row+i})'
    cell = ws.cell(row=egr_table_row + i, column=egr_table_col+1, value=f)
    cell.number_format = CURRENCY_BS
    cell.border = BORDER

bar = BarChart()
bar.title = "Egresos por Categoria"
data = Reference(ws, min_col=egr_table_col+1, min_row=egr_table_row, max_row=egr_table_row+len(egr_cats_list)-1)
cats_ref = Reference(ws, min_col=egr_table_col, min_row=egr_table_row, max_row=egr_table_row+len(egr_cats_list)-1)
bar.add_data(data)
bar.set_categories(cats_ref)
bar.height = 10
bar.width = 14
ws.add_chart(bar, "L" + str(chart_row))

# Cumplimiento presupuestario placeholder (semaforo simple via condicional)
sem_row = egr_table_row + len(egr_cats_list) + 2
ws.cell(row=sem_row, column=1, value="SEMAFORO DE LIQUIDEZ").font = H2_FONT
ws.cell(row=sem_row+1, column=1, value="Liquidez actual")
liq_cell = ws.cell(row=sem_row+1, column=2, value="=B12")  # referencia a KPI Liquidez
liq_cell.number_format = '0.00'
sem_cell = ws.cell(row=sem_row+1, column=3,
                    value='=IF(B'+str(sem_row+1)+'>=1.5,"VERDE",IF(B'+str(sem_row+1)+'>=1,"AMARILLO","ROJO"))')
sem_cell.font = LABEL_FONT

for col_letter, w in {"A": 18, "B": 16, "C": 12, "D": 4, "E": 4, "F": 4, "G": 4,
                       "H": 4, "I": 18, "J": 16, "K": 4, "L": 4}.items():
    ws.column_dimensions[col_letter].width = w

# ---------------------------------------------------------------
# 10_BALANCE_GENERAL
# ---------------------------------------------------------------
ws = wb.create_sheet("10_BALANCE_GENERAL")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:C1")
ws["A1"] = "BALANCE GENERAL"
style_header_row(ws, 1, 3)

ws["A3"] = "ACTIVOS"
ws["A3"].font = H1_FONT
ws["A3"].fill = SUBHEADER_FILL
ws.merge_cells("A3:C3")

ws["A4"] = "Activos Corrientes"
ws["A4"].font = H2_FONT
activos_corr = [
    ("Caja y Bancos", "='08_FLUJO_CAJA'!E15", CURRENCY_BS),
    ("Cuentas por Cobrar", "=SUMIFS('11_CUENTAS_POR_COBRAR'!E:E,'11_CUENTAS_POR_COBRAR'!F:F,\"Pendiente\")", CURRENCY_BS),
    ("Inventario", "='13_INVENTARIO'!F100", CURRENCY_BS),
]
r = 5
for label, formula, fmt in activos_corr:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=formula)
    cell.number_format = fmt
    cell.border = BORDER
    r += 1
ac_start, ac_end = 5, r - 1
ws.cell(row=r, column=1, value="Total Activos Corrientes").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=SUM(B{ac_start}:B{ac_end})")
cell.number_format = CURRENCY_BS
cell.font = LABEL_FONT
cell.fill = LIGHT_FILL
total_ac_row = r
r += 2

ws.cell(row=r, column=1, value="Activos No Corrientes").font = H2_FONT
r += 1
anc = [
    ("Propiedad, Planta y Equipo", 800000),
    ("Depreciacion Acumulada", -120000),
]
anc_start = r
for label, val in anc:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=val)
    cell.number_format = CURRENCY_BS
    cell.border = BORDER
    r += 1
anc_end = r - 1
ws.cell(row=r, column=1, value="Total Activos No Corrientes").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=SUM(B{anc_start}:B{anc_end})")
cell.number_format = CURRENCY_BS
cell.font = LABEL_FONT
cell.fill = LIGHT_FILL
total_anc_row = r
r += 1
ws.cell(row=r, column=1, value="TOTAL ACTIVOS").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=B{total_ac_row}+B{total_anc_row}")
cell.number_format = CURRENCY_BS
cell.font = LABEL_FONT
cell.fill = HEADER_FILL
cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
total_activos_row = r
r += 2

ws.cell(row=r, column=1, value="PASIVOS").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
r += 1
ws.cell(row=r, column=1, value="Pasivos Corrientes").font = H2_FONT
r += 1
pc = [
    ("Cuentas por Pagar", "=SUMIFS('12_CUENTAS_POR_PAGAR'!D:D,'12_CUENTAS_POR_PAGAR'!E:E,\"Pendiente\")"),
    ("Impuestos por Pagar", 25000),
    ("Prestamos a Corto Plazo", 50000),
]
pc_start = r
for label, val in pc:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=val)
    cell.number_format = CURRENCY_BS
    cell.border = BORDER
    r += 1
pc_end = r - 1
ws.cell(row=r, column=1, value="Total Pasivos Corrientes").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=SUM(B{pc_start}:B{pc_end})")
cell.number_format = CURRENCY_BS
cell.font = LABEL_FONT
cell.fill = LIGHT_FILL
total_pc_row = r
r += 2

ws.cell(row=r, column=1, value="Pasivos No Corrientes").font = H2_FONT
r += 1
pnc_start = r
ws.cell(row=r, column=1, value="Prestamos a Largo Plazo").font = NORMAL_FONT
cell = ws.cell(row=r, column=2, value=200000)
cell.number_format = CURRENCY_BS
cell.border = BORDER
pnc_end = r
r += 1
ws.cell(row=r, column=1, value="Total Pasivos No Corrientes").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=SUM(B{pnc_start}:B{pnc_end})")
cell.number_format = CURRENCY_BS
cell.font = LABEL_FONT
cell.fill = LIGHT_FILL
total_pnc_row = r
r += 1
ws.cell(row=r, column=1, value="TOTAL PASIVOS").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=B{total_pc_row}+B{total_pnc_row}")
cell.number_format = CURRENCY_BS
cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
cell.fill = HEADER_FILL
total_pasivos_row = r
r += 2

ws.cell(row=r, column=1, value="PATRIMONIO").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
r += 1
patr = [
    ("Capital Social", 300000),
    ("Utilidades Retenidas", 100000),
]
patr_start = r
for label, val in patr:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=val)
    cell.number_format = CURRENCY_BS
    cell.border = BORDER
    r += 1
ws.cell(row=r, column=1, value="Utilidad del Ejercicio").font = NORMAL_FONT
cell = ws.cell(row=r, column=2, value="=Utilidad_Neta")
cell.number_format = CURRENCY_BS
cell.border = BORDER
patr_end = r
r += 1
ws.cell(row=r, column=1, value="TOTAL PATRIMONIO").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=SUM(B{patr_start}:B{patr_end})")
cell.number_format = CURRENCY_BS
cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
cell.fill = HEADER_FILL
total_patrimonio_row = r
r += 2
ws.cell(row=r, column=1, value="TOTAL PASIVO + PATRIMONIO").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=B{total_pasivos_row}+B{total_patrimonio_row}")
cell.number_format = CURRENCY_BS
cell.font = LABEL_FONT
cell.fill = LIGHT_FILL
r += 2

ws.cell(row=r, column=1, value="INDICADORES FINANCIEROS").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
r += 1
ind = [
    ("Liquidez (Activo Cte. / Pasivo Cte.)", f"=B{total_ac_row}/B{total_pc_row}", '0.00'),
    ("Solvencia (Total Activos / Total Pasivos)", f"=B{total_activos_row}/B{total_pasivos_row}", '0.00'),
    ("Endeudamiento (Total Pasivos / Total Activos)", f"=B{total_pasivos_row}/B{total_activos_row}", PCT),
]
for label, formula, fmt in ind:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=formula)
    cell.number_format = fmt
    cell.border = BORDER
    cell.fill = LIGHT_FILL
    r += 1

autosize(ws, {"A": 38, "B": 18, "C": 4})

# ---------------------------------------------------------------
# 11_CUENTAS_POR_COBRAR
# ---------------------------------------------------------------
ws = wb.create_sheet("11_CUENTAS_POR_COBRAR")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:G1")
ws["A1"] = "CUENTAS POR COBRAR"
style_header_row(ws, 1, 7)

headers = ["Cliente", "Factura", "Fecha Emision", "Fecha Vencimiento", "Monto (Bs.)", "Estado", "Semaforo"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = H1_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER

cxc_data = [
    ("Cliente A", "FAC-001", date(2026, 5, 1), date(2026, 6, 1), 45000, "Pendiente"),
    ("Cliente B", "FAC-002", date(2026, 5, 10), date(2026, 6, 10), 18000, "Pendiente"),
    ("Cliente C", "FAC-003", date(2026, 4, 15), date(2026, 5, 15), 120000, "Pendiente"),
    ("Cliente D", "FAC-004", date(2026, 5, 20), date(2026, 6, 20), 5000, "Pagado"),
]
for i, row in enumerate(cxc_data, start=4):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=val)
        cell.border = BORDER
        if c in (3, 4):
            cell.number_format = "dd/mm/yyyy"
        if c == 5:
            cell.number_format = CURRENCY_BS
    sem = ws.cell(row=i, column=7,
                   value=(f'=IF(F{i}="Pagado","VERDE",'
                          f'IF(D{i}<TODAY(),"ROJO",'
                          f'IF(D{i}<=TODAY()+15,"AMARILLO","VERDE")))'))
    sem.alignment = CENTER
    sem.border = BORDER

cxc_last = 3 + len(cxc_data)
tbl = Table(displayName="tbl_CxC", ref=f"A3:G{cxc_last}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)

autosize(ws, {"A": 16, "B": 12, "C": 16, "D": 16, "E": 16, "F": 14, "G": 12})

# ---------------------------------------------------------------
# 12_CUENTAS_POR_PAGAR
# ---------------------------------------------------------------
ws = wb.create_sheet("12_CUENTAS_POR_PAGAR")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:F1")
ws["A1"] = "CUENTAS POR PAGAR"
style_header_row(ws, 1, 6)

headers = ["Proveedor / Obligacion", "Tipo", "Fecha Vencimiento", "Monto (Bs.)", "Estado", "Alerta"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = H1_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER

cxp_data = [
    ("Proveedor Insumos", "Compras", date(2026, 6, 20), 45000, "Pendiente"),
    ("SENIAT - IVA", "Impuestos", date(2026, 6, 15), 12000, "Pendiente"),
    ("Banco - Cuota Prestamo", "Creditos", date(2026, 6, 30), 8000, "Pendiente"),
    ("Taller Mecanico", "Compras", date(2026, 5, 1), 9000, "Pagado"),
]
for i, row in enumerate(cxp_data, start=4):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=val)
        cell.border = BORDER
        if c == 3:
            cell.number_format = "dd/mm/yyyy"
        if c == 4:
            cell.number_format = CURRENCY_BS
    alerta = ws.cell(row=i, column=6,
                      value=(f'=IF(E{i}="Pagado","Al dia",'
                             f'IF(C{i}<TODAY(),"VENCIDO",'
                             f'IF(C{i}<=TODAY()+7,"PROXIMO A VENCER","Al dia")))'))
    alerta.border = BORDER

cxp_last = 3 + len(cxp_data)
tbl = Table(displayName="tbl_CxP", ref=f"A3:F{cxp_last}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)

autosize(ws, {"A": 26, "B": 14, "C": 16, "D": 16, "E": 14, "F": 18})

# ---------------------------------------------------------------
# 13_INVENTARIO
# ---------------------------------------------------------------
ws = wb.create_sheet("13_INVENTARIO")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:H1")
ws["A1"] = "CONTROL DE INVENTARIO"
style_header_row(ws, 1, 8)

headers = ["Producto", "Existencia", "Stock Minimo", "Stock Maximo", "Costo Unitario (Bs.)",
           "Valor Inventario (Bs.)", "Rotacion (veces/mes)", "Cobertura (dias)"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = H1_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER

inv_data = [
    ("Producto X", 500, 100, 1000, 80, 6, 30),
    ("Producto Y", 200, 50, 500, 120, 4, 22),
    ("Producto Z", 80, 100, 600, 150, 2, 15),
    ("Materia Prima A", 1200, 300, 3000, 25, 8, 28),
]
for i, row in enumerate(inv_data, start=4):
    prod, exist, smin, smax, costo, rot, cob = row
    ws.cell(row=i, column=1, value=prod).border = BORDER
    ws.cell(row=i, column=2, value=exist).border = BORDER
    ws.cell(row=i, column=3, value=smin).border = BORDER
    ws.cell(row=i, column=4, value=smax).border = BORDER
    cell = ws.cell(row=i, column=5, value=costo); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=i, column=6, value=f"=B{i}*E{i}"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=i, column=7, value=rot); cell.border = BORDER
    cell = ws.cell(row=i, column=8, value=cob); cell.border = BORDER

inv_last = 3 + len(inv_data)
tbl = Table(displayName="tbl_Inventario", ref=f"A3:H{inv_last}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)

# Total valor inventario en celda F100 (referenciado por Balance General)
ws["E99"] = "VALOR TOTAL INVENTARIO"
ws["E99"].font = LABEL_FONT
cell = ws.cell(row=100, column=6, value=f"=SUM(F4:F{inv_last})")
cell.number_format = CURRENCY_BS
cell.font = LABEL_FONT
cell.fill = LIGHT_FILL

autosize(ws, {"A": 18, "B": 12, "C": 12, "D": 12, "E": 18, "F": 18, "G": 18, "H": 16})

# ---------------------------------------------------------------
# 14_NOMINA
# ---------------------------------------------------------------
ws = wb.create_sheet("14_NOMINA")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:K1")
ws["A1"] = "NOMINA - LEGISLACION LABORAL VENEZOLANA (LOTTT, IVSS, INCES, FAOV)"
style_header_row(ws, 1, 11)

headers = ["Empleado", "Cargo", "Salario Mensual (Bs.)", "Dias Trabajados",
           "IVSS Trabajador (4%)", "FAOV Trabajador (1%)", "Aporte Trabajador",
           "IVSS Patrono (9%)", "FAOV Patrono (2%)", "INCES Patrono (2%)", "Salario Neto (Bs.)"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = H1_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER

nom_data = [
    ("Empleado 1", "Gerente Administrativo", 15000, 30),
    ("Empleado 2", "Analista Financiero", 9000, 30),
    ("Empleado 3", "Operador de Planta", 6000, 30),
    ("Empleado 4", "Conductor", 5500, 30),
]
for i, (nombre, cargo, salario, dias) in enumerate(nom_data, start=4):
    ws.cell(row=i, column=1, value=nombre).border = BORDER
    ws.cell(row=i, column=2, value=cargo).border = BORDER
    cell = ws.cell(row=i, column=3, value=salario); cell.number_format = CURRENCY_BS; cell.border = BORDER
    ws.cell(row=i, column=4, value=dias).border = BORDER
    cell = ws.cell(row=i, column=5, value=f"=C{i}*0.04"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=i, column=6, value=f"=C{i}*0.01"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=i, column=7, value=f"=E{i}+F{i}"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=i, column=8, value=f"=C{i}*0.09"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=i, column=9, value=f"=C{i}*0.02"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=i, column=10, value=f"=C{i}*0.02"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=i, column=11, value=f"=C{i}-G{i}"); cell.number_format = CURRENCY_BS; cell.border = BORDER

nom_last = 3 + len(nom_data)
tbl = Table(displayName="tbl_Nomina", ref=f"A3:K{nom_last}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)

# Totales
tot_row = nom_last + 1
ws.cell(row=tot_row, column=2, value="TOTALES").font = LABEL_FONT
for col in [3, 5, 6, 7, 8, 9, 10, 11]:
    cl = get_column_letter(col)
    cell = ws.cell(row=tot_row, column=col, value=f"=SUM({cl}4:{cl}{nom_last})")
    cell.number_format = CURRENCY_BS
    cell.font = LABEL_FONT
    cell.fill = LIGHT_FILL

# Calculo de Prestaciones, Vacaciones y Utilidades (anual, simplificado)
r = tot_row + 3
ws.cell(row=r, column=1, value="PRESTACIONES SOCIALES, VACACIONES Y UTILIDADES (Estimacion Anual)").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
headers2 = ["Empleado", "Salario Mensual (Bs.)", "Prestaciones (15 dias x 12 meses)",
            "Vacaciones (15 dias)", "Utilidades (30 dias min. LOTTT)"]
for i, h in enumerate(headers2, start=1):
    cell = ws.cell(row=r, column=i, value=h)
    cell.font = H1_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
r += 1
for i, (nombre, cargo, salario, dias) in enumerate(nom_data):
    row_ref = 4 + i
    ws.cell(row=r, column=1, value=nombre).border = BORDER
    cell = ws.cell(row=r, column=2, value=f"=C{row_ref}"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=r, column=3, value=f"=(C{row_ref}/30)*15*12"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=r, column=4, value=f"=(C{row_ref}/30)*15"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=r, column=5, value=f"=(C{row_ref}/30)*30"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    r += 1

autosize(ws, {"A": 16, "B": 22, "C": 18, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 16, "K": 16})

# ---------------------------------------------------------------
# 15_IMPUESTOS
# ---------------------------------------------------------------
ws = wb.create_sheet("15_IMPUESTOS")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:D1")
ws["A1"] = "CONTROL TRIBUTARIO - SENIAT Y MUNICIPAL"
style_header_row(ws, 1, 4)

ws["A3"] = "SENIAT - IVA"
ws["A3"].font = H1_FONT
ws["A3"].fill = SUBHEADER_FILL
ws.merge_cells("A3:B3")

iva_rows = [
    ("IVA Debito Fiscal (Ventas x Alicuota)", f"=Ingresos_Totales*'02_CONFIGURACION'!$B$6", CURRENCY_BS),
    ("IVA Credito Fiscal (Compras x Alicuota)",
     f'=SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"Compras")*\'02_CONFIGURACION\'!$B$6', CURRENCY_BS),
    ("IVA a Pagar (Debito - Credito)", "=B4-B5", CURRENCY_BS),
]
r = 4
for label, formula, fmt in iva_rows:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=formula)
    cell.number_format = fmt
    cell.border = BORDER
    r += 1

r += 1
ws.cell(row=r, column=1, value="RETENCIONES").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
ret_rows = [
    ("Retencion IVA (75% sobre IVA causado)", "=B5*0.75", CURRENCY_BS),
    ("Retencion ISLR (estimado 2% sobre compras)",
     '=SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"Compras")*0.02', CURRENCY_BS),
]
for label, formula, fmt in ret_rows:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=formula)
    cell.number_format = fmt
    cell.border = BORDER
    r += 1

r += 1
ws.cell(row=r, column=1, value="IGTF (Pagos en divisas)").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
ws.cell(row=r, column=1, value="IGTF Estimado (Egresos USD x Alicuota)").font = NORMAL_FONT
cell = ws.cell(row=r, column=2, value=f"=Egresos_Totales/Tasa_BCV*'02_CONFIGURACION'!$B$7")
cell.number_format = CURRENCY_USD
cell.border = BORDER
r += 2

ws.cell(row=r, column=1, value="IMPUESTOS MUNICIPALES").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
mun_rows = [
    ("Impuesto Actividades Economicas (1% s/Ingresos)", "=Ingresos_Totales*0.01", CURRENCY_BS),
    ("Impuesto Publicidad y Propaganda", 1500, CURRENCY_BS),
    ("Impuesto Inmuebles Urbanos", 2500, CURRENCY_BS),
]
for label, formula, fmt in mun_rows:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=formula)
    cell.number_format = fmt
    cell.border = BORDER
    r += 1

autosize(ws, {"A": 42, "B": 18, "C": 4, "D": 4})

# ---------------------------------------------------------------
# 16_PRESUPUESTO
# ---------------------------------------------------------------
ws = wb.create_sheet("16_PRESUPUESTO")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:E1")
ws["A1"] = "PRESUPUESTO: PRESUPUESTADO VS EJECUTADO"
style_header_row(ws, 1, 5)

headers = ["Concepto", "Presupuestado (Bs.)", "Ejecutado (Bs.)", "Variacion Absoluta (Bs.)", "Variacion (%)"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = H1_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER

presupuesto_rows = [
    ("Ingresos", 700000, "=Ingresos_Totales"),
    ("Costo de Ventas", 180000, '=SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"Compras")'),
    ("Gastos Operativos", 250000,
     '=SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"Nomina")'
     '+SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"Combustible")'
     '+SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"Transporte")'
     '+SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"Mantenimiento")'
     '+SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"Servicios")'
     '+SUMIFS(tbl_Egresos[Bs.],tbl_Egresos[Categoria],"Otros gastos")'),
    ("Gastos Financieros", 15000, "=Gastos_Financieros"),
]
r = 4
for label, presup, ejec in presupuesto_rows:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=presup); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=r, column=3, value=ejec); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=r, column=4, value=f"=C{r}-B{r}"); cell.number_format = CURRENCY_BS; cell.border = BORDER
    cell = ws.cell(row=r, column=5, value=f"=D{r}/B{r}"); cell.number_format = PCT; cell.border = BORDER
    r += 1

autosize(ws, {"A": 24, "B": 18, "C": 18, "D": 20, "E": 14})

# ---------------------------------------------------------------
# 17_KPI_GERENCIALES
# ---------------------------------------------------------------
ws = wb.create_sheet("17_KPI_GERENCIALES")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:C1")
ws["A1"] = "KPI GERENCIALES"
style_header_row(ws, 1, 3)

ws["A3"] = "FINANCIEROS"
ws["A3"].font = H1_FONT
ws["A3"].fill = SUBHEADER_FILL
ws.merge_cells("A3:B3")
fin_kpis = [
    ("EBITDA (Bs.)", "=Utilidad_Operativa", CURRENCY_BS),
    ("ROI (Utilidad Neta / Total Activos)", f"=Utilidad_Neta/'10_BALANCE_GENERAL'!B{total_activos_row}", PCT),
    ("ROA (Utilidad Operativa / Total Activos)", f"=Utilidad_Operativa/'10_BALANCE_GENERAL'!B{total_activos_row}", PCT),
    ("ROE (Utilidad Neta / Patrimonio)", f"=Utilidad_Neta/'10_BALANCE_GENERAL'!B{total_patrimonio_row}", PCT),
    ("Margen Neto", "=Utilidad_Neta/Ingresos_Totales", PCT),
    ("Liquidez", f"='10_BALANCE_GENERAL'!B{total_ac_row}/'10_BALANCE_GENERAL'!B{total_pc_row}", '0.00'),
]
r = 4
for label, formula, fmt in fin_kpis:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=formula)
    cell.number_format = fmt
    cell.border = BORDER
    cell.fill = LIGHT_FILL
    r += 1

r += 1
ws.cell(row=r, column=1, value="COMERCIALES").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
com_kpis = [
    ("Crecimiento de Ventas (vs presupuesto)", "='16_PRESUPUESTO'!E4", PCT),
    ("Clientes Activos (con CxC)", "=COUNTA(tbl_CxC[Cliente])", '0'),
]
for label, formula, fmt in com_kpis:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=formula)
    cell.number_format = fmt
    cell.border = BORDER
    cell.fill = LIGHT_FILL
    r += 1

r += 1
ws.cell(row=r, column=1, value="OPERATIVOS").font = H1_FONT
ws.cell(row=r, column=1).fill = SUBHEADER_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
op_kpis = [
    ("Productividad (Ingresos / Nro. Empleados)", f"=Ingresos_Totales/{len(nom_data)}", CURRENCY_BS),
    ("Eficiencia (Utilidad Operativa / Ingresos)", "=Utilidad_Operativa/Ingresos_Totales", PCT),
]
for label, formula, fmt in op_kpis:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=formula)
    cell.number_format = fmt
    cell.border = BORDER
    cell.fill = LIGHT_FILL
    r += 1

autosize(ws, {"A": 40, "B": 18, "C": 4})

# ---------------------------------------------------------------
# MODULO MINERO
# ---------------------------------------------------------------

# 18_PRODUCCION_MINERA
ws = wb.create_sheet("18_PRODUCCION_MINERA")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:E1")
ws["A1"] = "PRODUCCION MINERA"
style_header_row(ws, 1, 5)

headers = ["Fecha", "Mina", "Mineral", "Produccion (Tm)", "Ventas (Tm)"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = H1_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER

prod_data = [
    (date(2026, 1, 31), "Mina 1", "Oro", 120, 110),
    (date(2026, 2, 28), "Mina 1", "Oro", 135, 130),
    (date(2026, 3, 31), "Mina 1", "Oro", 140, 138),
    (date(2026, 4, 30), "Mina 2", "Hierro", 5000, 4800),
    (date(2026, 5, 31), "Mina 2", "Hierro", 5200, 5100),
]
for i, (f, mina, mineral, prod, vent) in enumerate(prod_data, start=4):
    ws.cell(row=i, column=1, value=f).number_format = "dd/mm/yyyy"
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=2, value=mina).border = BORDER
    ws.cell(row=i, column=3, value=mineral).border = BORDER
    ws.cell(row=i, column=4, value=prod).border = BORDER
    ws.cell(row=i, column=5, value=vent).border = BORDER
    cell = ws.cell(row=i, column=6, value=f"=SUM(D$4:D{i})-SUM(E$4:E{i})")
    cell.border = BORDER
ws.cell(row=3, column=6, value="Inventario (Tm)").font = H1_FONT
ws.cell(row=3, column=6).fill = SUBHEADER_FILL
ws.cell(row=3, column=6).alignment = CENTER

prod_last = 3 + len(prod_data)
tbl = Table(displayName="tbl_ProduccionMinera", ref=f"A3:F{prod_last}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)

autosize(ws, {"A": 14, "B": 12, "C": 12, "D": 16, "E": 16, "F": 16})

# 19_COSTO_POR_TONELADA
ws = wb.create_sheet("19_COSTO_POR_TONELADA")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:C1")
ws["A1"] = "COSTO POR TONELADA"
style_header_row(ws, 1, 3)

cpt_rows = [
    ("Costo de Extraccion (Bs./Tm)", 35),
    ("Costo de Carga (Bs./Tm)", 8),
    ("Costo de Transporte (Bs./Tm)", 12),
    ("Costo de Beneficio (Bs./Tm)", 20),
    ("Costo de Comercializacion (Bs./Tm)", 5),
]
r = 3
cpt_start = r
for label, val in cpt_rows:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=val)
    cell.number_format = CURRENCY_BS
    cell.border = BORDER
    r += 1
cpt_end = r - 1
ws.cell(row=r, column=1, value="COSTO TOTAL POR TONELADA").font = LABEL_FONT
cell = ws.cell(row=r, column=2, value=f"=SUM(B{cpt_start}:B{cpt_end})")
cell.number_format = CURRENCY_BS
cell.font = LABEL_FONT
cell.fill = LIGHT_FILL
costo_tm_row = r
add_named("Costo_Total_Tonelada", f"='19_COSTO_POR_TONELADA'!B{costo_tm_row}")

autosize(ws, {"A": 32, "B": 16, "C": 4})

# 20_RESERVAS_MINERAS
ws = wb.create_sheet("20_RESERVAS_MINERAS")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:E1")
ws["A1"] = "RESERVAS MINERAS"
style_header_row(ws, 1, 5)

headers = ["Mina", "Reservas Medidas (Tm)", "Reservas Indicadas (Tm)", "Reservas Inferidas (Tm)", "Vida Util (anos)"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = H1_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER

res_data = [
    ("Mina 1", 50000, 30000, 20000),
    ("Mina 2", 800000, 500000, 300000),
]
for i, (mina, med, ind, inf) in enumerate(res_data, start=4):
    ws.cell(row=i, column=1, value=mina).border = BORDER
    ws.cell(row=i, column=2, value=med).number_format = '#,##0'
    ws.cell(row=i, column=2).border = BORDER
    ws.cell(row=i, column=3, value=ind).number_format = '#,##0'
    ws.cell(row=i, column=3).border = BORDER
    ws.cell(row=i, column=4, value=inf).number_format = '#,##0'
    ws.cell(row=i, column=4).border = BORDER
    # Vida util = reservas medidas / produccion anual estimada de esa mina
    cell = ws.cell(row=i, column=5,
                    value=f'=B{i}/SUMPRODUCT((tbl_ProduccionMinera[Mina]=A{i})*tbl_ProduccionMinera[Produccion (Tm)])*1')
    cell.number_format = '0.0'
    cell.border = BORDER

res_last = 3 + len(res_data)
tbl = Table(displayName="tbl_Reservas", ref=f"A3:E{res_last}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)

autosize(ws, {"A": 14, "B": 22, "C": 22, "D": 22, "E": 16})

# 21_INDICADORES_MINEROS
ws = wb.create_sheet("21_INDICADORES_MINEROS")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:C1")
ws["A1"] = "INDICADORES MINEROS"
style_header_row(ws, 1, 3)

min_kpis = [
    ("Produccion Mensual (ultimo periodo, Tm)", "=INDEX(tbl_ProduccionMinera[Produccion (Tm)],COUNTA(tbl_ProduccionMinera[Mina]))", '#,##0'),
    ("Produccion Anual (Tm)", "=SUM(tbl_ProduccionMinera[Produccion (Tm)])", '#,##0'),
    ("Cumplimiento Plan de Explotacion (%)", "=B5/600000", PCT),
    ("Productividad por Trabajador (Tm/trabajador)", f"=B5/{len(nom_data)}", '#,##0.0'),
    ("Productividad por Equipo (Tm/equipo, 5 equipos)", "=B5/5", '#,##0.0'),
    ("Costo por Tonelada (Bs.)", "=Costo_Total_Tonelada", CURRENCY_BS),
    ("Utilidad por Tonelada (Bs.)", "=Precio_Venta_Unitario-Costo_Total_Tonelada", CURRENCY_BS),
    ("Indice de Accidentabilidad (accidentes/100 trab.)", 0, '0.0'),
    ("Cumplimiento Ambiental (%)", 0.95, PCT),
]
r = 4
for label, formula, fmt in min_kpis:
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    cell = ws.cell(row=r, column=2, value=formula)
    cell.number_format = fmt
    cell.border = BORDER
    cell.fill = LIGHT_FILL
    r += 1

autosize(ws, {"A": 44, "B": 18, "C": 4})

# ---------------------------------------------------------------
# 00_MANUAL (usuario + tecnico + migracion a Power BI)
# ---------------------------------------------------------------
ws = wb.create_sheet("00_MANUAL")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:B1")
ws["A1"] = "MANUAL DE USO Y MANUAL TECNICO"
style_header_row(ws, 1, 2)

manual_text = [
    ("MANUAL DE USUARIO", True),
    ("1. Navegacion", False),
    ("Use los enlaces de la hoja 01_PORTADA o las pestanas inferiores para moverse entre modulos.", False),
    ("2. Configuracion inicial", False),
    ("Antes de usar el archivo, complete 02_CONFIGURACION: datos de la empresa, RIF, tasa BCV,", False),
    ("alicuotas de IVA/IGTF, centros de costos, lineas de negocio y categorias.", False),
    ("3. Registro de operaciones", False),
    ("Ingrese cada transaccion en 04_INGRESOS y 05_EGRESOS dentro de las tablas (tbl_Ingresos /", False),
    ("tbl_Egresos). Al escribir una nueva fila debajo de la tabla, esta se expande automaticamente", False),
    ("y todas las formulas, dashboards y reportes se recalculan solos.", False),
    ("4. Categorias", False),
    ("Use unicamente las categorias definidas en 02_CONFIGURACION para que las formulas SUMIFS", False),
    ("del Estado de Resultados, Presupuesto e Impuestos funcionen correctamente.", False),
    ("5. Cuentas por Cobrar / Pagar", False),
    ("La columna 'Semaforo' / 'Alerta' se calcula sola segun la fecha de vencimiento y el estado", False),
    ("(Pendiente/Pagado). Verde = al dia, Amarillo = proximo a vencer, Rojo = vencido.", False),
    ("6. Actualizar el Dashboard", False),
    ("El Dashboard Ejecutivo se recalcula automaticamente (formulas nativas de Excel). Si usa", False),
    ("Power Query, presione 'Actualizar todo' en la pestana Datos.", False),
    ("", False),
    ("MANUAL TECNICO", True),
    ("1. Estructura de nombres definidos", False),
    ("Tasa_BCV, Costos_Fijos, Costo_Var_Unitario, Precio_Venta_Unitario, Unidades_Vendidas,", False),
    ("Ingresos_Totales, Utilidad_Bruta, Utilidad_Operativa, Utilidad_Neta, Gastos_Financieros,", False),
    ("Egresos_Totales, Costo_Total_Tonelada (ver Formulas > Administrador de nombres).", False),
    ("2. Tablas estructuradas (Excel Tables)", False),
    ("tbl_Ingresos, tbl_Egresos, tbl_CxC, tbl_CxP, tbl_Inventario, tbl_Nomina,", False),
    ("tbl_ProduccionMinera, tbl_Reservas. No elimine ni renombre estas tablas: las formulas de", False),
    ("Dashboard, Estado de Resultados, Flujo de Caja, Presupuesto e Indicadores dependen de ellas.", False),
    ("3. Relaciones entre hojas", False),
    ("04/05 -> 09 (Estado Resultados, via SUMIFS) -> 10 (Balance, Utilidad Neta)", False),
    ("04/05 -> 08 (Flujo de Caja, via SUMPRODUCT por mes/ano)", False),
    ("11/12/13 -> 10 (Balance General: CxC, CxP, Inventario)", False),
    ("06 -> 07 (Punto de Equilibrio, via nombres definidos)", False),
    ("18 -> 19, 20, 21 (Modulo minero)", False),
    ("4. Mantenimiento", False),
    ("Para agregar un nuevo centro de costos, linea de negocio o categoria, agreguela en", False),
    ("02_CONFIGURACION y, si corresponde, ajuste las formulas SUMIFS en 09, 15, 16 y 21.", False),
    ("5. Proteccion", False),
    ("Se recomienda proteger las hojas con formulas (Revisar > Proteger hoja) dejando", False),
    ("desbloqueadas unicamente las celdas de captura (dentro de las tablas tbl_*).", False),
    ("6. Automatizacion VBA", False),
    ("Importe el modulo 'ERP_Macros.bas' (incluido junto a este archivo) en el Editor de VBA", False),
    ("(Alt+F11 > Importar archivo). Incluye macros para actualizar todas las tablas dinamicas", False),
    ("y consultas de Power Query, y para exportar el Dashboard a PDF.", False),
    ("", False),
    ("MIGRACION FUTURA A POWER BI / DASHBOARD WEB", True),
    ("1. Conectar Power BI Desktop a este archivo (Obtener datos > Excel) y seleccionar las", False),
    ("tablas tbl_Ingresos, tbl_Egresos, tbl_CxC, tbl_CxP, tbl_Inventario, tbl_Nomina,", False),
    ("tbl_ProduccionMinera y tbl_Reservas.", False),
    ("2. Crear un modelo en estrella: tablas de hechos (Ingresos, Egresos, Produccion) y", False),
    ("dimensiones (Calendario, Categorias, Centros de Costos, Clientes/Proveedores, Minas).", False),
    ("3. Replicar las medidas DAX equivalentes a las formulas de 09_ESTADO_RESULTADOS,", False),
    ("07_PUNTO_EQUILIBRIO y 19_COSTO_POR_TONELADA usando CALCULATE/SUMX.", False),
    ("4. Publicar el modelo en Power BI Service y configurar actualizacion programada apuntando", False),
    ("al archivo Excel en OneDrive/SharePoint.", False),
    ("5. Para un dashboard web propio, exponer las mismas consultas mediante una API", False),
    ("(Power Automate / Azure Functions) que lea el libro desde SharePoint y alimente un", False),
    ("frontend (Power Apps o aplicacion web a medida).", False),
]

r = 3
for text, is_title in manual_text:
    cell = ws.cell(row=r, column=1, value=text)
    if is_title:
        cell.font = H1_FONT
        cell.fill = SUBHEADER_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    else:
        cell.font = NORMAL_FONT
    r += 1

autosize(ws, {"A": 100, "B": 4})

# ---------------------------------------------------------------
# Orden final de hojas
# ---------------------------------------------------------------
order = ["01_PORTADA", "00_MANUAL", "02_CONFIGURACION", "03_DASHBOARD_EJECUTIVO", "04_INGRESOS",
         "05_EGRESOS", "06_COSTOS", "07_PUNTO_EQUILIBRIO", "08_FLUJO_CAJA", "09_ESTADO_RESULTADOS",
         "10_BALANCE_GENERAL", "11_CUENTAS_POR_COBRAR", "12_CUENTAS_POR_PAGAR", "13_INVENTARIO",
         "14_NOMINA", "15_IMPUESTOS", "16_PRESUPUESTO", "17_KPI_GERENCIALES",
         "18_PRODUCCION_MINERA", "19_COSTO_POR_TONELADA", "20_RESERVAS_MINERAS", "21_INDICADORES_MINEROS"]
wb._sheets = [wb[name] for name in order]

wb.active = 0

out_path = r"C:\Users\FUNDESTA\ObservatorioMinero\Sistema_Gestion_Financiera.xlsx"
wb.save(out_path)
print("Archivo generado:", out_path)
